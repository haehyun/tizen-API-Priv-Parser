from openpyxl import Workbook


class getInstance:


    def __init__(self):
        self.wb = Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = "public"
        self.ws2 = self.wb.create_sheet()
        self.ws2.title = "platform"
        self.ws3 = self.wb.create_sheet()
        self.ws3.title = "partner"

        self.public_count = 1
        self.platform_count = 1
        self.parter_count = 1


    def writeAPI(self, api):
        privlevel = api.getPrivlevel()
        privilege = api.getPrivilege()
        name = api.getName()


        if privlevel == "public":
            self.ws1.cell(row=self.public_count, column=2).value = str(privilege)
            self.ws1.cell(row=self.public_count, column=1).value = name
            self.public_count += 1

        elif privlevel == "platform":
            self.ws2.cell(row=self.platform_count, column=2).value = str(privilege)
            self.ws2.cell(row=self.platform_count, column=1).value = name
            self.platform_count += 1

        else:
            self.ws3.cell(row=self.parter_count, column=2).value = str(privilege)
            self.ws3.cell(row=self.parter_count, column=1).value = name
            self.parter_count += 1


    def save(self, path):
        self.wb.save(path + "/cc.xlsx")
