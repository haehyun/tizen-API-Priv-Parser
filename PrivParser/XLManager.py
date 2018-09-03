from openpyxl import Workbook


class getInstance:


    def __init__(self):
        self.wb = Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = "functions"
        self.ws2 = self.wb.create_sheet()
        self.ws2.title = "privileges"
        # self.ws3 = self.wb.create_sheet()
        # self.ws3.title = "privileges"

        self.functions_count = 1
        self.privileges_count = 1
        # self.parter_count = 1


    def readXL(self, title):
        if title == "func":
            for row in self.ws1.rows:
                for cell in row:
                    print(cell.value)
        elif title == "priv":
            for row in self.ws2.rows:
                for cell in row:
                    print(cell.value)    

    def writeXL(self, input, size, title):
        if title == "func":
            for x in range(0, size):
                self.ws1.cell(row=self.functions_count, column=1).value = input[x]
                # self.ws1.cell(row=self.functions, column=2).value = str(privilege)
                self.functions_count += 1
        elif title == "priv":
            for x in range(0, size):
                self.ws2.cell(row=self.functions_count, column=1).value = input[x]
                # self.ws2.cell(row=self.privileges, column=2).value = str(privilege)
                self.privileges_count += 1


    def save(self, path):
        self.wb.save(path + "/DA_output.xlsx")